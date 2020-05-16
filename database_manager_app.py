# This program is intended to manage excel workbooks and handle data in excel
from openpyxl import load_workbook
from openpyxl import Workbook
import pickle
from tqdm import *
import time

# This is the welcome message to make this app seem more professional
print("Welcome to the database manager app! Enjoy yourself as this program does all the repetitive task for you.\nThis program is specifically made to tailor to your needs by Captain Candy™\n\n")

# This is a dictionary that specifies how many groups every class has to be assigned
rows_to_move = {
                'Nur' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'Kg' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '1' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '2' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '3' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '4' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '5' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '6' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '7' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '8' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '9' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '10' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '11' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                '12' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0}
}

# This part asks you if you want to change the data present in the rows_to_move dictonary.
i  = 0
while i == 0:
    intention_to_change = input("Do you want to change the number of rows that each class is moved? (Only answer in y or n)\n")
    if intention_to_change == 'y':
        break
    elif intention_to_change == 'n':
        break
    else:
        print("\n!Please type in a valid character!\n")

# These are the conditions which fulfill your request. If data has to be changed, then it will allow you to change it. If data does not have to changed then it will load the previous saved data.
if intention_to_change == 'y':
    print("I would now like for you to type in the number of groups that have to be assigned to each class.")
    NurA = input('Nursury A\n')
    NurB = input('Nursury B\n')
    NurC = input('Nursury C\n')
    NurD = input('Nursury D\n')
    NurE = input('Nursury E\n')
    NurF = input('Nursury F\n')
    NurG = input('Nursury G\n')
    NurH = input('Nursury H\n')
    NurI = input('Nursury I\n')
    KgA = input('Kindergarten A\n')
    KgB = input('Kindergarten B\n')
    KgC = input('Kindergarten C\n')
    KgD = input('Kindergarten D\n')
    KgE = input('Kindergarten E\n')
    KgF = input('Kindergarten F\n')
    KgG = input('Kindergarten G\n')
    KgH = input('Kindergarten H\n')
    KgI = input('Kindergarten I\n')
    FirstA = input('First A\n')
    FirstB = input('First B\n')
    FirstC = input('First C\n')
    FirstD = input('First D\n')
    FirstE = input('First E\n')
    FirstF = input('First F\n')
    FirstG = input('First G\n')
    FirstH = input('First H\n')
    FirstI = input('First I\n')
    SecondA = input('Second A\n')
    SecondB = input('Second B\n')
    SecondC = input('Second C\n')
    SecondD = input('Second D\n')
    SecondE = input('Second E\n')
    SecondF = input('Second F\n')
    SecondG = input('Second G\n')
    SecondH = input('Second H\n')
    SecondI = input('Second I\n')
    ThirdA = input('Third A\n')
    ThirdB = input('Third B\n')
    ThirdC = input('Third C\n')
    ThirdD = input('Third D\n')
    ThirdE = input('Third E\n')
    ThirdF = input('Third F\n')
    ThirdG = input('Third G\n')
    ThirdH = input('Third H\n')
    ThirdI = input('Third I\n')
    FourthA = input('Fourth A\n')
    FourthB = input('Fourth B\n')
    FourthC = input('Fourth C\n')
    FourthD = input('Fourth D\n')
    FourthE = input('Fourth E\n')
    FourthF = input('Fourth F\n')
    FourthG = input('Fourth G\n')
    FourthH = input('Fourth H\n')
    FourthI = input('Fourth I\n')
    FifthA = input('Fifth A\n')
    FifthB = input('Fifth B\n')
    FifthC = input('Fifth C\n')
    FifthD = input('Fifth D\n')
    FifthE = input('Fifth E\n')
    FifthF = input('Fifth F\n')
    FifthG = input('Fifth G\n')
    FifthH = input('Fifth H\n')
    FifthI = input('Fifth I\n')
    SixthA = input('Sixth A\n')
    SixthB = input('Sixth B\n')
    SixthC = input('Sixth C\n')
    SixthD = input('Sixth D\n')
    SixthE = input('Sixth E\n')
    SixthF = input('Sixth F\n')
    SixthG = input('Sixth G\n')
    SixthH = input('Sixth H\n')
    SixthI = input('Sixth I\n')
    SeventhA = input('Seventh A\n')
    SeventhB = input('Seventh B\n')
    SeventhC = input('Seventh C\n')
    SeventhD = input('Seventh D\n')
    SeventhE = input('Seventh E\n')
    SeventhF = input('Seventh F\n')
    SeventhG = input('Seventh G\n')
    SeventhH = input('Seventh H\n')
    SeventhI = input('Seventh I\n')
    EighthA = input('Eighth A\n')
    EighthB = input('Eighth B\n')
    EighthC = input('Eighth C\n')
    EighthD = input('Eighth D\n')
    EighthE = input('Eighth E\n')
    EighthF = input('Eighth F\n')
    EighthG = input('Eighth G\n')
    EighthH = input('Eighth H\n')
    EighthI = input('Eighth I\n')
    NinthA = input('Ninth A\n')
    NinthB = input('Ninth B\n')
    NinthC = input('Ninth C\n')
    NinthD = input('Ninth D\n')
    NinthE = input('Ninth E\n')
    NinthF = input('Ninth F\n')
    NinthG = input('Ninth G\n')
    NinthH = input('Ninth H\n')
    NinthI = input('Ninth I\n')
    TenthA = input('Tenth A\n')
    TenthB = input('Tenth B\n')
    TenthC = input('Tenth C\n')
    TenthD = input('Tenth D\n')
    TenthE = input('Tenth E\n')
    TenthF = input('Tenth F\n')
    TenthG = input('Tenth G\n')
    TenthH = input('Tenth H\n')
    TenthI = input('Tenth I\n')
    EleventhA = input('Eleventh A\n')
    EleventhB = input('Eleventh B\n')
    EleventhC = input('Eleventh C\n')
    EleventhD = input('Eleventh D\n')
    EleventhE = input('Eleventh E\n')
    EleventhF = input('Eleventh F\n')
    EleventhG = input('Eleventh G\n')
    EleventhH = input('Eleventh H\n')
    EleventhI = input('Eleventh I\n')
    TwelfthA = input('Twelfth A\n')
    TwelfthB = input('Twelfth B\n')
    TwelfthC = input('Twelfth C\n')
    TwelfthD = input('Twelfth D\n')
    TwelfthE = input('Twelfth E\n')
    TwelfthF = input('Twelfth F\n')
    TwelfthG = input('Twelfth G\n')
    TwelfthH = input('Twelfth H\n')
    TwelfthI = input('Twelfth I\n')

    rows_to_move['Nur']['a'] = NurA
    rows_to_move['Nur']['b'] = NurB
    rows_to_move['Nur']['c'] = NurC
    rows_to_move['Nur']['d'] = NurD
    rows_to_move['Nur']['e'] = NurE
    rows_to_move['Nur']['f'] = NurF
    rows_to_move['Nur']['g'] = NurG
    rows_to_move['Nur']['h'] = NurH
    rows_to_move['Nur']['i'] = NurI
    rows_to_move['Kg']['a'] = KgA
    rows_to_move['Kg']['b'] = KgB
    rows_to_move['Kg']['c'] = KgC
    rows_to_move['Kg']['d'] = KgD
    rows_to_move['Kg']['e'] = KgE
    rows_to_move['Kg']['f'] = KgF
    rows_to_move['Kg']['g'] = KgG
    rows_to_move['Kg']['h'] = KgH
    rows_to_move['Kg']['i'] = KgI
    rows_to_move['1']['a'] = FirstA
    rows_to_move['1']['b'] = FirstB
    rows_to_move['1']['c'] = FirstC
    rows_to_move['1']['d'] = FirstD
    rows_to_move['1']['e'] = FirstE
    rows_to_move['1']['f'] = FirstF
    rows_to_move['1']['g'] = FirstG
    rows_to_move['1']['h'] = FirstH
    rows_to_move['1']['i'] = FirstI
    rows_to_move['2']['a'] = SecondA
    rows_to_move['2']['b'] = SecondB
    rows_to_move['2']['c'] = SecondC
    rows_to_move['2']['d'] = SecondD
    rows_to_move['2']['e'] = SecondE
    rows_to_move['2']['f'] = SecondF
    rows_to_move['2']['g'] = SecondG
    rows_to_move['2']['h'] = SecondH
    rows_to_move['2']['i'] = SecondI
    rows_to_move['3']['a'] = ThirdA
    rows_to_move['3']['b'] = ThirdB
    rows_to_move['3']['c'] = ThirdC
    rows_to_move['3']['d'] = ThirdD
    rows_to_move['3']['e'] = ThirdE
    rows_to_move['3']['f'] = ThirdF
    rows_to_move['3']['g'] = ThirdG
    rows_to_move['3']['h'] = ThirdH
    rows_to_move['3']['i'] = ThirdI
    rows_to_move['4']['a'] = FourthA
    rows_to_move['4']['b'] = FourthB
    rows_to_move['4']['c'] = FourthC
    rows_to_move['4']['d'] = FourthD
    rows_to_move['4']['e'] = FourthE
    rows_to_move['4']['f'] = FourthF
    rows_to_move['4']['g'] = FourthG
    rows_to_move['4']['h'] = FourthH
    rows_to_move['4']['i'] = FourthI
    rows_to_move['5']['a'] = FifthA
    rows_to_move['5']['b'] = FifthB
    rows_to_move['5']['c'] = FifthC
    rows_to_move['5']['d'] = FifthD
    rows_to_move['5']['e'] = FifthE
    rows_to_move['5']['f'] = FifthF
    rows_to_move['5']['g'] = FifthG
    rows_to_move['5']['h'] = FifthH
    rows_to_move['5']['i'] = FifthI
    rows_to_move['6']['a'] = SixthA
    rows_to_move['6']['b'] = SixthB
    rows_to_move['6']['c'] = SixthC
    rows_to_move['6']['d'] = SixthD
    rows_to_move['6']['e'] = SixthE
    rows_to_move['6']['f'] = SixthF
    rows_to_move['6']['g'] = SixthG
    rows_to_move['6']['h'] = SixthH
    rows_to_move['6']['i'] = SixthI
    rows_to_move['7']['a'] = SeventhA
    rows_to_move['7']['b'] = SeventhB
    rows_to_move['7']['c'] = SeventhC
    rows_to_move['7']['d'] = SeventhD
    rows_to_move['7']['e'] = SeventhE
    rows_to_move['7']['f'] = SeventhF
    rows_to_move['7']['g'] = SeventhG
    rows_to_move['7']['h'] = SeventhH
    rows_to_move['7']['i'] = SeventhI
    rows_to_move['8']['a'] = EighthA
    rows_to_move['8']['b'] = EighthB
    rows_to_move['8']['c'] = EighthC
    rows_to_move['8']['d'] = EighthD
    rows_to_move['8']['e'] = EighthE
    rows_to_move['8']['f'] = EighthF
    rows_to_move['8']['g'] = EighthG
    rows_to_move['8']['h'] = EighthH
    rows_to_move['8']['i'] = EighthI
    rows_to_move['9']['a'] = NinthA
    rows_to_move['9']['b'] = NinthB
    rows_to_move['9']['c'] = NinthC
    rows_to_move['9']['d'] = NinthD
    rows_to_move['9']['e'] = NinthE
    rows_to_move['9']['f'] = NinthF
    rows_to_move['9']['g'] = NinthG
    rows_to_move['9']['h'] = NinthH
    rows_to_move['9']['i'] = NinthI
    rows_to_move['10']['a'] = TenthA
    rows_to_move['10']['b'] = TenthB
    rows_to_move['10']['c'] = TenthC
    rows_to_move['10']['d'] = TenthD
    rows_to_move['10']['e'] = TenthE
    rows_to_move['10']['f'] = TenthF
    rows_to_move['10']['g'] = TenthG
    rows_to_move['10']['h'] = TenthH
    rows_to_move['10']['i'] = TenthI
    rows_to_move['11']['a'] = EleventhA
    rows_to_move['11']['b'] = EleventhB
    rows_to_move['11']['c'] = EleventhC
    rows_to_move['11']['d'] = EleventhD
    rows_to_move['11']['e'] = EleventhE
    rows_to_move['11']['f'] = EleventhF
    rows_to_move['11']['g'] = EleventhG
    rows_to_move['11']['h'] = EleventhH
    rows_to_move['11']['i'] = EleventhI
    rows_to_move['12']['a'] = TwelfthA
    rows_to_move['12']['b'] = TwelfthB
    rows_to_move['12']['c'] = TwelfthC
    rows_to_move['12']['d'] = TwelfthD
    rows_to_move['12']['e'] = TwelfthE
    rows_to_move['12']['f'] = TwelfthF
    rows_to_move['12']['g'] = TwelfthG
    rows_to_move['12']['h'] = TwelfthH
    rows_to_move['12']['i'] = TwelfthI

    with open('data_pick.pkl', 'wb') as pickle_file:
        pickle.dump(rows_to_move, pickle_file)
    print("Repository successfully updated")

elif intention_to_change == 'n':
    with open('data_pick.pkl', 'rb') as pickle_file:
        rows_to_move = pickle.load(pickle_file)
    print("Great, the number of groups that will be assigned will be the data provided previously.\n...\nRepository successfully loaded")

# This part is where the user chooses which file they want to copy data from and which file it has to be sent to.
students = input('Which file contains the data of the Students? \n')
groups = input('Which file contains the data of the Groups? \n')
print("Processing...")

# This part is defining the openpyxl objects for later use.
wb = load_workbook(students + ".xlsx")
wb2 = load_workbook(groups + ".xlsx")
final_wb = Workbook()
sheet = wb.active
sheet2 = wb2.active
final_sheet = final_wb.active

# This is the loop which repeats untill all the data has been copied into the other file.
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        final_sheet.cell(row = i, column = j).value = sheet.cell(row = i, column = j).value
print("Student data copied!")

# This part defines some variables that are to be used in the later code.
letter_array = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
letter_array_small = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
last_column_letter = letter_array[final_sheet.max_column]
last_row_number = str(sheet.max_row)
columns_to_move = 0
class_array = ['Nursury','Kg','I','II','III','IV','V','VI','VII','VIII','IX','X','XI','XII']
caMod = 0
directory_array = ['Nur','Kg','1','2','3','4','5','6','7','8','9','10','11','12']
daMod = 0
i = 2
l = 1

print("Starting to space out data...")

# This loop spaces out the rows on the basis of the number of groups which have to assigned to each kid.
while i <= sheet.max_row*13:
    if final_sheet.cell(row = i, column = 10).value == class_array[caMod]:
        for l in range(1, 10):
            if final_sheet.cell(row = i, column = 11).value  == letter_array[l-1]:
                rows_moved = (int(rows_to_move[directory_array[daMod]][letter_array_small[l-1]])-1)
                final_sheet.move_range(f"A{i+1}:{last_column_letter+last_row_number}", rows=rows_moved, cols=columns_to_move)
                i = i + (rows_moved+1)
                last_row_number = str(sheet.max_row+i)
                break
            else:
                l = l + 1

    elif final_sheet.cell(row = i, column = 10).value == '':
        break
    else:
        caMod = caMod + 1
        daMod = daMod + 1
        if caMod == 14:
            break

print("Data spaced out!")

# Here both the files that were used in the process are saved as seperate files so as to not corrupt the original data in case of an accident.
wb.save(students + "2.xlsx")
wb2.save(groups + "2.xlsx")
final_wb.save("Final_Data.xlsx")
final_wb.close()
wb.close()
wb2.close()

# This part fills in the gaps that were created during spacing out.
final_wb = load_workbook("Final_Data.xlsx")
final_sheet  = final_wb.active
i = 2
l = 1
caMod = 0
daMod = 0

pbar = tqdm(total = final_sheet.max_row+10)

while i < final_sheet.max_row+13:
    if final_sheet.cell(row = i, column = 10).value == class_array[caMod]:
        for l in range(1, 10):
            if final_sheet.cell(row = i, column = 11).value  == letter_array[l-1]:
                rows_moved = (int(rows_to_move[directory_array[daMod]][letter_array_small[l-1]]))
                for x in range(1, rows_moved):
                    for y in range(1, final_sheet.max_column+1):
                        final_sheet.cell(row = i+x, column = y).value = final_sheet.cell(row = i, column = y).value
                    pbar.update(1)
                i = i + (rows_moved)
                break
            else:
                l = l + 1

    elif final_sheet.cell(row = i, column = 10).value == '':
        break
    else:
        caMod = caMod + 1
        daMod = daMod + 1
        if caMod == 14:
            break

pbar.close()
final_wb.save("Final_Data.xlsx")
final_wb.close()

# This part assigns the 'All' classes
final_wb = load_workbook("Final_Data.xlsx")
final_sheet  = final_wb.active
wb2 = load_workbook(groups + ".xlsx")
sheet2 = wb2.active

i = 2
l = 1
caMod = 0
daMod = 0

pbar = tqdm(total = final_sheet.max_row/11)

while i < final_sheet.max_row+13:
    if final_sheet.cell(row = i, column = 10).value == class_array[caMod]:
        rows_moved = (int(rows_to_move[directory_array[daMod]][letter_array_small[l-1]]))
        for r in range(1, 15):
            if sheet2.cell(row = r, column = 6).value == 'All':
                for y in range(1, sheet2.max_column+1):
                    final_sheet.cell(row = i, column = y+final_sheet.max_column+1).value = sheet2.cell(row = 6, column = y).value
                pbar.update(1)
        i = i + (rows_moved)

    elif final_sheet.cell(row = i, column = 10).value == '':
        break
    else:
        caMod = caMod + 1
        daMod = daMod + 1
        if caMod == 14:
            break

pbar.close()
time.sleep(2)
final_wb.save("Final_Data.xlsx")

# This part tells python that is is done using the file and that it can close them and get then off the RAM.
wb2.save(groups + "2.xlsx")
wb2.close()
final_wb.close()

# This is just to tell the user that what they have done has succeded and now they can live a stressfree life.
print('task completed\nHave a good day\n\n')
close_the_program_please = input("Press enter to close the program")