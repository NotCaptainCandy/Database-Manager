from openpyxl import load_workbook
import pickle
from tqdm import *

rows_to_move = {
                'Nursery' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'KG' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'I' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'II' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'III' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'IV' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'V' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'VI' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'VII' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'VIII' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'IX' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'X' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'XI' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0},
                'XII' : {'a' : 0,
                        'b' : 0,
                        'c' : 0,
                        'd' : 0,
                        'e' : 0,
                        'f' : 0,
                        'g' : 0,
                        'h' : 0,
                        'i' : 0}
}

i  = 0
while i == 0:
    intention_to_change = input("Do you want to change the number of rows that each class is moved? (Only answer in y or n)\n")
    if intention_to_change == 'y':
        break
    elif intention_to_change == 'n':
        break
    else:
        print("\n!Please type in a valid character!\n")

if intention_to_change == 'y':
    print("I would now like for you to type in the number of groups that have to be assigned to each class.")
    NurA = input('Nursury A\n')
    NurB = input('Nursury B\n')
    NurC = input('Nursury C\n')
    NurD = input('Nursury D\n')
    NurE = input('Nursury E\n')
    NurF = input('Nursury F\n')
    NurG = input('Nursury G\n')
    NurH = input('Nursury H\n')
    NurI = input('Nursury I\n')
    KgA = input('Kindergarten A\n')
    KgB = input('Kindergarten B\n')
    KgC = input('Kindergarten C\n')
    KgD = input('Kindergarten D\n')
    KgE = input('Kindergarten E\n')
    KgF = input('Kindergarten F\n')
    KgG = input('Kindergarten G\n')
    KgH = input('Kindergarten H\n')
    KgI = input('Kindergarten I\n')
    FirstA = input('First A\n')
    FirstB = input('First B\n')
    FirstC = input('First C\n')
    FirstD = input('First D\n')
    FirstE = input('First E\n')
    FirstF = input('First F\n')
    FirstG = input('First G\n')
    FirstH = input('First H\n')
    FirstI = input('First I\n')
    SecondA = input('Second A\n')
    SecondB = input('Second B\n')
    SecondC = input('Second C\n')
    SecondD = input('Second D\n')
    SecondE = input('Second E\n')
    SecondF = input('Second F\n')
    SecondG = input('Second G\n')
    SecondH = input('Second H\n')
    SecondI = input('Second I\n')
    ThirdA = input('Third A\n')
    ThirdB = input('Third B\n')
    ThirdC = input('Third C\n')
    ThirdD = input('Third D\n')
    ThirdE = input('Third E\n')
    ThirdF = input('Third F\n')
    ThirdG = input('Third G\n')
    ThirdH = input('Third H\n')
    ThirdI = input('Third I\n')
    FourthA = input('Fourth A\n')
    FourthB = input('Fourth B\n')
    FourthC = input('Fourth C\n')
    FourthD = input('Fourth D\n')
    FourthE = input('Fourth E\n')
    FourthF = input('Fourth F\n')
    FourthG = input('Fourth G\n')
    FourthH = input('Fourth H\n')
    FourthI = input('Fourth I\n')
    FifthA = input('Fifth A\n')
    FifthB = input('Fifth B\n')
    FifthC = input('Fifth C\n')
    FifthD = input('Fifth D\n')
    FifthE = input('Fifth E\n')
    FifthF = input('Fifth F\n')
    FifthG = input('Fifth G\n')
    FifthH = input('Fifth H\n')
    FifthI = input('Fifth I\n')
    SixthA = input('Sixth A\n')
    SixthB = input('Sixth B\n')
    SixthC = input('Sixth C\n')
    SixthD = input('Sixth D\n')
    SixthE = input('Sixth E\n')
    SixthF = input('Sixth F\n')
    SixthG = input('Sixth G\n')
    SixthH = input('Sixth H\n')
    SixthI = input('Sixth I\n')
    SeventhA = input('Seventh A\n')
    SeventhB = input('Seventh B\n')
    SeventhC = input('Seventh C\n')
    SeventhD = input('Seventh D\n')
    SeventhE = input('Seventh E\n')
    SeventhF = input('Seventh F\n')
    SeventhG = input('Seventh G\n')
    SeventhH = input('Seventh H\n')
    SeventhI = input('Seventh I\n')
    EighthA = input('Eighth A\n')
    EighthB = input('Eighth B\n')
    EighthC = input('Eighth C\n')
    EighthD = input('Eighth D\n')
    EighthE = input('Eighth E\n')
    EighthF = input('Eighth F\n')
    EighthG = input('Eighth G\n')
    EighthH = input('Eighth H\n')
    EighthI = input('Eighth I\n')
    NinthA = input('Ninth A\n')
    NinthB = input('Ninth B\n')
    NinthC = input('Ninth C\n')
    NinthD = input('Ninth D\n')
    NinthE = input('Ninth E\n')
    NinthF = input('Ninth F\n')
    NinthG = input('Ninth G\n')
    NinthH = input('Ninth H\n')
    NinthI = input('Ninth I\n')
    TenthA = input('Tenth A\n')
    TenthB = input('Tenth B\n')
    TenthC = input('Tenth C\n')
    TenthD = input('Tenth D\n')
    TenthE = input('Tenth E\n')
    TenthF = input('Tenth F\n')
    TenthG = input('Tenth G\n')
    TenthH = input('Tenth H\n')
    TenthI = input('Tenth I\n')
    EleventhA = input('Eleventh A\n')
    EleventhB = input('Eleventh B\n')
    EleventhC = input('Eleventh C\n')
    EleventhD = input('Eleventh D\n')
    EleventhE = input('Eleventh E\n')
    EleventhF = input('Eleventh F\n')
    EleventhG = input('Eleventh G\n')
    EleventhH = input('Eleventh H\n')
    EleventhI = input('Eleventh I\n')
    TwelfthA = input('Twelfth A\n')
    TwelfthB = input('Twelfth B\n')
    TwelfthC = input('Twelfth C\n')
    TwelfthD = input('Twelfth D\n')
    TwelfthE = input('Twelfth E\n')
    TwelfthF = input('Twelfth F\n')
    TwelfthG = input('Twelfth G\n')
    TwelfthH = input('Twelfth H\n')
    TwelfthI = input('Twelfth I\n')

    rows_to_move['Nursery']['a'] = NurA
    rows_to_move['Nursery']['b'] = NurB
    rows_to_move['Nursery']['c'] = NurC
    rows_to_move['Nursery']['d'] = NurD
    rows_to_move['Nursery']['e'] = NurE
    rows_to_move['Nursery']['f'] = NurF
    rows_to_move['Nursery']['g'] = NurG
    rows_to_move['Nursery']['h'] = NurH
    rows_to_move['Nursery']['i'] = NurI
    rows_to_move['KG']['a'] = KgA
    rows_to_move['KG']['b'] = KgB
    rows_to_move['KG']['c'] = KgC
    rows_to_move['KG']['d'] = KgD
    rows_to_move['KG']['e'] = KgE
    rows_to_move['KG']['f'] = KgF
    rows_to_move['KG']['g'] = KgG
    rows_to_move['KG']['h'] = KgH
    rows_to_move['KG']['i'] = KgI
    rows_to_move['I']['a'] = FirstA
    rows_to_move['I']['b'] = FirstB
    rows_to_move['I']['c'] = FirstC
    rows_to_move['I']['d'] = FirstD
    rows_to_move['I']['e'] = FirstE
    rows_to_move['I']['f'] = FirstF
    rows_to_move['I']['g'] = FirstG
    rows_to_move['I']['h'] = FirstH
    rows_to_move['I']['i'] = FirstI
    rows_to_move['II']['a'] = SecondA
    rows_to_move['II']['b'] = SecondB
    rows_to_move['II']['c'] = SecondC
    rows_to_move['II']['d'] = SecondD
    rows_to_move['II']['e'] = SecondE
    rows_to_move['II']['f'] = SecondF
    rows_to_move['II']['g'] = SecondG
    rows_to_move['II']['h'] = SecondH
    rows_to_move['II']['i'] = SecondI
    rows_to_move['III']['a'] = ThirdA
    rows_to_move['III']['b'] = ThirdB
    rows_to_move['III']['c'] = ThirdC
    rows_to_move['III']['d'] = ThirdD
    rows_to_move['III']['e'] = ThirdE
    rows_to_move['III']['f'] = ThirdF
    rows_to_move['III']['g'] = ThirdG
    rows_to_move['III']['h'] = ThirdH
    rows_to_move['III']['i'] = ThirdI
    rows_to_move['IV']['a'] = FourthA
    rows_to_move['IV']['b'] = FourthB
    rows_to_move['IV']['c'] = FourthC
    rows_to_move['IV']['d'] = FourthD
    rows_to_move['IV']['e'] = FourthE
    rows_to_move['IV']['f'] = FourthF
    rows_to_move['IV']['g'] = FourthG
    rows_to_move['IV']['h'] = FourthH
    rows_to_move['IV']['i'] = FourthI
    rows_to_move['V']['a'] = FifthA
    rows_to_move['V']['b'] = FifthB
    rows_to_move['V']['c'] = FifthC
    rows_to_move['V']['d'] = FifthD
    rows_to_move['V']['e'] = FifthE
    rows_to_move['V']['f'] = FifthF
    rows_to_move['V']['g'] = FifthG
    rows_to_move['V']['h'] = FifthH
    rows_to_move['V']['i'] = FifthI
    rows_to_move['VI']['a'] = SixthA
    rows_to_move['VI']['b'] = SixthB
    rows_to_move['VI']['c'] = SixthC
    rows_to_move['VI']['d'] = SixthD
    rows_to_move['VI']['e'] = SixthE
    rows_to_move['VI']['f'] = SixthF
    rows_to_move['VI']['g'] = SixthG
    rows_to_move['VI']['h'] = SixthH
    rows_to_move['VI']['i'] = SixthI
    rows_to_move['VII']['a'] = SeventhA
    rows_to_move['VII']['b'] = SeventhB
    rows_to_move['VII']['c'] = SeventhC
    rows_to_move['VII']['d'] = SeventhD
    rows_to_move['VII']['e'] = SeventhE
    rows_to_move['VII']['f'] = SeventhF
    rows_to_move['VII']['g'] = SeventhG
    rows_to_move['VII']['h'] = SeventhH
    rows_to_move['VII']['i'] = SeventhI
    rows_to_move['VIII']['a'] = EighthA
    rows_to_move['VIII']['b'] = EighthB
    rows_to_move['VIII']['c'] = EighthC
    rows_to_move['VIII']['d'] = EighthD
    rows_to_move['VIII']['e'] = EighthE
    rows_to_move['VIII']['f'] = EighthF
    rows_to_move['VIII']['g'] = EighthG
    rows_to_move['VIII']['h'] = EighthH
    rows_to_move['VIII']['i'] = EighthI
    rows_to_move['IX']['a'] = NinthA
    rows_to_move['IX']['b'] = NinthB
    rows_to_move['IX']['c'] = NinthC
    rows_to_move['IX']['d'] = NinthD
    rows_to_move['IX']['e'] = NinthE
    rows_to_move['IX']['f'] = NinthF
    rows_to_move['IX']['g'] = NinthG
    rows_to_move['IX']['h'] = NinthH
    rows_to_move['IX']['i'] = NinthI
    rows_to_move['X']['a'] = TenthA
    rows_to_move['X']['b'] = TenthB
    rows_to_move['X']['c'] = TenthC
    rows_to_move['X']['d'] = TenthD
    rows_to_move['X']['e'] = TenthE
    rows_to_move['X']['f'] = TenthF
    rows_to_move['X']['g'] = TenthG
    rows_to_move['X']['h'] = TenthH
    rows_to_move['X']['i'] = TenthI
    rows_to_move['XI']['a'] = EleventhA
    rows_to_move['XI']['b'] = EleventhB
    rows_to_move['XI']['c'] = EleventhC
    rows_to_move['XI']['d'] = EleventhD
    rows_to_move['XI']['e'] = EleventhE
    rows_to_move['XI']['f'] = EleventhF
    rows_to_move['XI']['g'] = EleventhG
    rows_to_move['XI']['h'] = EleventhH
    rows_to_move['XI']['i'] = EleventhI
    rows_to_move['XII']['a'] = TwelfthA
    rows_to_move['XII']['b'] = TwelfthB
    rows_to_move['XII']['c'] = TwelfthC
    rows_to_move['XII']['d'] = TwelfthD
    rows_to_move['XII']['e'] = TwelfthE
    rows_to_move['XII']['f'] = TwelfthF
    rows_to_move['XII']['g'] = TwelfthG
    rows_to_move['XII']['h'] = TwelfthH
    rows_to_move['XII']['i'] = TwelfthI

    with open('data_pick.pkl', 'wb') as pickle_file:
        pickle.dump(rows_to_move, pickle_file)
    print("Repository successfully updated")

elif intention_to_change == 'n':
    with open('data_pick.pkl', 'rb') as pickle_file:
        rows_to_move = pickle.load(pickle_file)
    print("Great, the number of groups that will be assigned will be the data provided previously.\n...\nRepository successfully loaded")

final_wb = load_workbook("Final_Data.xlsx")
final_sheet  = final_wb.active
wb2 = load_workbook("groups.xlsx")
sheet2 = wb2.active

letter_array = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
letter_array_small = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
last_column_letter = letter_array[final_sheet.max_column]
class_array = ['Nursery','KG','I','II','III','IV','V','VI','VII','VIII','IX','X','XI','XII']
caMod = 0
directory_array = ['Nursery','KG','I','II','III','IV','V','VI','VII','VIII','IX','X','XI','XII']
daMod = 0
l = 1
i = 2
final_max_column = final_sheet.max_column
rows_moved = (int(rows_to_move[directory_array[daMod]][letter_array_small[l-1]]))

for i in range(1,8):
    final_sheet.cell(row=1, column=i+final_max_column).value = sheet2.cell(row = 1, column=i).value

all_files = [[0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0],
             [0,0,0,0,0,0,0,0]]

for i in range(1 , sheet2.max_row):
    if sheet2.cell(row = i, column = 6).value == "All":
        for j in range(14):
            if sheet2.cell(row = i, column = 5).value == class_array[j]:
                for z in range(8):
                    all_files[j][z] = sheet2.cell(row = i, column= z+1).value
                break

i = 2

while i < final_sheet.max_row:
    while True:
        if final_sheet.cell(row=i, column=10).value == class_array[caMod]:
            for j in range(1,8):
                final_sheet.cell(row=i, column=j+final_max_column).value = all_files[caMod][j-1]
            rows_moved = (int(rows_to_move[directory_array[daMod]][letter_array_small[l-1]]))
            i = i + rows_moved
            break
        else:
            caMod = caMod + 1
            daMod = daMod + 1

# This part is going to assign the groups to the students based on their classes.

groups_data = []

for i in range(2, sheet2.max_row):
    row_data = []
    for x in range(1,9):
        row_data.append(sheet2.cell(row = i, column = x).value)
    groups_data.append(row_data)

print("Groups data successfully read\nBeginning to assign groups to students")

i = 3
g = 0

pbar = tqdm(total = final_sheet.max_row)

while i < final_sheet.max_row:
    if final_sheet.cell(row = i, column = 10).value == groups_data[g][4]:
        if final_sheet.cell(row = i, column = 11).value == groups_data[g][5]:
            rows_moved = (int(rows_to_move[groups_data[g][4]][str(groups_data[g][5]).lower()])-1)
            for p in range(rows_moved):
                pbar.update(1)
                for col in range(8):
                    final_sheet.cell(row = i+p, column = 14+col).value = groups_data[g+p][col]
            i = i + rows_moved+1
        else:
            g = g + 1
    else:
        g = g + 1

pbar.close()

# Finish

final_wb.save("Final_Data2.xlsx")
final_wb.close()
wb2.close()

